import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import threading

class ExcelWriter:
    def __init__(self, filename='glaucoma_test_results.xlsx'):
        self.filename = filename
        self.lock = threading.Lock()
        self.initialize_workbook()
    
    def initialize_workbook(self):
        """Initialize Excel workbook with test sheets"""
        try:
            if os.path.exists(self.filename):
                self.workbook = load_workbook(self.filename)
            else:
                self.workbook = Workbook()
                # Remove default sheet
                if 'Sheet' in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook['Sheet'])
            
            # Create sheets for each test type
            test_sheets = [
                'Visual Field', 'CSV-1000', 'Edge Detection', 
                'Motion Detection', 'Pattern Recognition', 
                'Pelli-Robinson', 'SPARCS', 'Doctor Notes'
            ]
            
            for sheet_name in test_sheets:
                if sheet_name not in self.workbook.sheetnames:
                    sheet = self.workbook.create_sheet(sheet_name)
                    self.setup_sheet_headers(sheet, sheet_name)
            
            self.save_workbook()
            logging.info(f"Excel workbook initialized: {self.filename}")
            
        except Exception as e:
            logging.error(f"Error initializing workbook: {e}")
    
    def setup_sheet_headers(self, sheet, sheet_name):
        """Setup headers for each sheet type"""
        common_headers = ['Patient Name', 'Test Date', 'Start Time', 'End Time', 
                         'Duration (seconds)', 'Total Points', 'Correct Points', 
                         'Accuracy (%)', 'Doctor Notes']
        
        # Add test-specific headers
        if sheet_name == 'Visual Field':
            headers = common_headers + ['Points Tested', 'Sensitivity Map', 'Defects Detected']
        elif sheet_name == 'CSV-1000':
            headers = common_headers + ['Language', 'Contrast Levels', 'Letter Accuracy']
        elif sheet_name == 'Pelli-Robinson':
            headers = common_headers + ['Language', 'Contrast Sensitivity', 'Log Units']
        elif sheet_name == 'SPARCS':
            headers = common_headers + ['Quadrant 1', 'Quadrant 2', 'Quadrant 3', 'Quadrant 4']
        elif sheet_name == 'Doctor Notes':
            headers = ['Patient Name', 'Date', 'Symptoms', 'Medical Concerns', 'Additional Notes']
        else:
            headers = common_headers + ['Test Specific Data']
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    def save_workbook(self):
        """Save workbook to file"""
        try:
            self.workbook.save(self.filename)
        except Exception as e:
            logging.error(f"Error saving workbook: {e}")
    
    def save_test_result(self, test_name, patient_name, test_data):
        """Save test results to appropriate sheet"""
        with self.lock:
            try:
                sheet_mapping = {
                    'visual_field': 'Visual Field',
                    'csv1000': 'CSV-1000',
                    'edge': 'Edge Detection',
                    'motion': 'Motion Detection',
                    'pattern': 'Pattern Recognition',
                    'pelli_robinson': 'Pelli-Robinson',
                    'sparcs': 'SPARCS'
                }
                
                sheet_name = sheet_mapping.get(test_name, test_name)
                if sheet_name not in self.workbook.sheetnames:
                    sheet = self.workbook.create_sheet(sheet_name)
                    self.setup_sheet_headers(sheet, sheet_name)
                else:
                    sheet = self.workbook[sheet_name]
                
                # Find next empty row
                next_row = sheet.max_row + 1
                
                # Common data
                current_time = datetime.now()
                duration = test_data.get('duration', 0)
                total_points = test_data.get('total_points', 0)
                correct_points = test_data.get('correct_points', 0)
                accuracy = (correct_points / total_points * 100) if total_points > 0 else 0
                
                # Base row data
                row_data = [
                    patient_name,
                    current_time.strftime('%Y-%m-%d'),
                    test_data.get('start_time', current_time.strftime('%H:%M:%S')),
                    test_data.get('end_time', current_time.strftime('%H:%M:%S')),
                    duration,
                    total_points,
                    correct_points,
                    round(accuracy, 2),
                    test_data.get('doctor_notes', '')
                ]
                
                # Add test-specific data
                if test_name == 'visual_field':
                    row_data.extend([
                        test_data.get('points_tested', 0),
                        str(test_data.get('sensitivity_map', [])),
                        test_data.get('defects_detected', 0)
                    ])
                elif test_name == 'csv1000':
                    row_data.extend([
                        test_data.get('language', 'English'),
                        str(test_data.get('contrast_levels', [])),
                        test_data.get('letter_accuracy', 0)
                    ])
                elif test_name == 'pelli_robinson':
                    row_data.extend([
                        test_data.get('language', 'English'),
                        test_data.get('contrast_sensitivity', 0),
                        test_data.get('log_units', 0)
                    ])
                elif test_name == 'sparcs':
                    row_data.extend([
                        test_data.get('quadrant_1', 0),
                        test_data.get('quadrant_2', 0),
                        test_data.get('quadrant_3', 0),
                        test_data.get('quadrant_4', 0)
                    ])
                else:
                    row_data.append(str(test_data.get('specific_data', '')))
                
                # Write row data
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=next_row, column=col, value=value)
                
                self.save_workbook()
                logging.info(f"Saved {test_name} results for patient {patient_name}")
                
            except Exception as e:
                logging.error(f"Error saving test result: {e}")
    
    def save_doctor_notes(self, patient_name, notes_data):
        """Save doctor's notes to dedicated sheet"""
        with self.lock:
            try:
                sheet_name = 'Doctor Notes'
                if sheet_name not in self.workbook.sheetnames:
                    sheet = self.workbook.create_sheet(sheet_name)
                    self.setup_sheet_headers(sheet, sheet_name)
                else:
                    sheet = self.workbook[sheet_name]
                
                # Find next empty row
                next_row = sheet.max_row + 1
                current_time = datetime.now()
                
                row_data = [
                    patient_name,
                    current_time.strftime('%Y-%m-%d %H:%M:%S'),
                    notes_data.get('symptoms', ''),
                    notes_data.get('medical_concerns', ''),
                    notes_data.get('additional_notes', '')
                ]
                
                # Write row data
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=next_row, column=col, value=value)
                
                self.save_workbook()
                logging.info(f"Saved doctor notes for patient {patient_name}")
                
            except Exception as e:
                logging.error(f"Error saving doctor notes: {e}")
    
    def get_patient_history(self, patient_name):
        """Get all test history for a patient"""
        history = {}
        try:
            for sheet_name in self.workbook.sheetnames:
                if sheet_name == 'Doctor Notes':
                    continue
                    
                sheet = self.workbook[sheet_name]
                patient_rows = []
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] == patient_name:  # Patient name is first column
                        patient_rows.append(row)
                
                if patient_rows:
                    history[sheet_name] = patient_rows
                    
        except Exception as e:
            logging.error(f"Error getting patient history: {e}")
        
        return history
